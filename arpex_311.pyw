"""modul ustvari/posodobi Excelovo datoteko z danimi podatki v .txt dokumentu"""
# https://www.youtube.com/watch?v=QWqxRchawZY -> standalone tkinter file
# https://openpyxl.readthedocs.io/en/stable/ - > openpyxl docs

import os
from os import sys

from tkinter.messagebox import askyesno
import tkinter.filedialog

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import GradientFill, Font


class Individual():
    """razred loci podatke v .txt datoteki;
    objekt je posameznik z atributom 1 vrstice .txt datoteke"""

    def __init__(self, string):
        self.string = string

    def get_name(self):
        """metoda poisce ime"""
        exception=0
        comma = self.string.find(", ,")
        if comma == -1:
            exception=1
        _name = self.string[:comma]
        return _name.title(), exception

    def get_school(self):
        """metoda poisce ime osnovne sole"""
        exception=0
        comma = self.string.find(", ,")
        bracket = self.string.find("(")
        if -1 in (comma,bracket):
            exception=1
        _school = self.string[comma+4:bracket-1]
        return _school, exception

    def get_mail(self):
        """metoda poisce e-naslov"""
        exception=0
        bracket_open = self.string.find("(")
        bracket_close = self.string.find(")")
        if -1 in (bracket_open,bracket_close):
            exception=1
        _mail = self.string[bracket_open+1:bracket_close]
        return _mail, exception

    def get_status(self):
        """metoda loci ucitelje od skrbnikov"""
        _status = ""
        if "učitelj" in self.string.lower():
            _status = "učitelj"
        if "skrbni" in self.string.lower():
            _status = "skrbnik"
        return _status

    def get_year(self):
        """metoda poisce leto"""
        dot1 = self.string.find(".")
        dot2 = (dot1+1) + self.string[dot1+1:].find(".")
        month = self.string[dot1+1:dot2]
        year = self.string[dot2+1:dot2+5]
        try:
            if int(month) < 9:
                previous_year = int(year)-1
                school_year = f"{previous_year}/{year}"
            else:
                next_year = int(year)+1
                school_year = f"{year}/{next_year}"
            return str(school_year)
        except ValueError:
            err_txt_format()
            return None
        # millenium = self.string.find("20")
        # _year = self.string[millenium:millenium+9]
        # return _year


def make_headings():
    """funkcija v xlsx datoteki ustvari naslove kategorij"""
    headings = ["IME&PRIIMEK", "ŠOLSKO LETO", "ŠOLA", "STATUS", "E-MAIL"]
    ws_participants.append(headings)
    return headings

def set_col_width(headings):
    """funkcija nastavi zacetno sirino stolpcev"""
    for col, header in enumerate(headings):
        col_letter = chr(64+col+1)
        ws_participants.column_dimensions[col_letter].width = len(header)+7

def read_txt(path):
    """funkcija prebere .txt datoteko s podatki"""
    with open(path, encoding="utf-8") as datasheet:
        for line in datasheet:
            if line in ("\n"," "):
                continue
            if "POKONČNA" in line:
                try:
                    obj = Individual(line)
                    status = obj.get_status()
                    year = obj.get_year()
                    continue
                except UnboundLocalError: # as error:
                    # print(error)
                    err_txt_format()
            split_data(line,status,year)

def split_data(line,status,year):
    """funkcija loci podatke v vrstici"""
    # klicanje metode z vrstico v podatkih kot lastnost posameznika (objekta)
    obj = Individual(line)
    name, exc1 = obj.get_name()
    school, exc2 = obj.get_school()
    mail, exc3 = obj.get_mail()
    exceptions = [exc1, exc2, exc3]
    year_status = (year, status)
    sort_data(name, school, mail, year_status, exceptions)

def sort_data(name, school, mail, year_status, exceptions):
    """funkcija sortira podatke"""
    year, status = year_status
    in_excel, name_row = in_xlsx(name)
    if in_excel:
        old_year = int((name_row[1].value)[:4])
        new_year = int(year[:4])
    if not in_excel:
        export_data(name, year, school, status, mail)
        row = last_row()
        if status == "skrbnik":
            color_row(row, "C4D79B")
        if 1 in exceptions:
            color_row(row, "FFFF66")
    elif new_year>old_year:
        data_string = [name, year, school, status, mail]
        update_data(name_row, data_string)

def export_data(name, year, school, status, mail):
    """funckija izvozi podatke v Excel, ce jih tam se ni"""
    data_string = [name, year, school, status, mail]
    for col_idx, value in enumerate(data_string, start=1):
        col = chr(64+col_idx)
        adjust_col_width(value, col)
    ws_participants.append(data_string)

def update_data(row, data_string):
    """funckija popravi stare podatke v Excelu"""
    for col_idx, value in enumerate(data_string, start=1):
        col = chr(64+col_idx)
        adjust_col_width(value, col)
        row[col_idx-1].value = value

def adjust_col_width(value, col):
    """funkcija prilagodi sirino stolpca"""
    width = ws_participants.column_dimensions[col].width
    if len(value) > width:
        ws_participants.column_dimensions[col].width = len(value)

def color_row(row, color):
    """fukcija pobarva dano vrstico"""
    for _row in ws_participants.iter_cols(min_col=1, max_col=5, min_row=row):
        for cell in _row:
            cell.fill = GradientFill(stop=(color, color))

def make_table():
    """funkcija v xlsx datoteki ustvari preglednico"""
    tab = Table(displayName="Udeleženci", ref=f"A1:{last_col()}{last_row()}")
    style = TableStyleInfo(name="TableStyleLight14", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws_participants.add_table(tab)
    set_font(font)

def set_font(_font):
    """funkcija doloci slog tabele"""
    for _row in ws_participants[f"A1:{last_col()}{last_row()}"]:
        for _cell in _row:
            _cell.font = _font

def in_xlsx(name):
    """funkcija preveri, ce so podatki ze v Excelovi datoteki"""
    row_ = None
    for _row in ws_participants.iter_rows(min_row=2, max_row=last_row()):
        row_ = _row
        if name == _row[0].value:
            return True, _row
    return False, row_


def last_row():
    """funkcija vrne zadnjo vrstico v Excelovi tabeli"""
    _last_row = ws_participants.max_row
    return _last_row

def last_col():
    """funkcija vrne zadnji stolpec v Excelovi tabeli"""
    last_col_index = ws_participants.max_column
    _last_col = chr(64+last_col_index)
    return _last_col

def create_xlsx():
    """seznam ukazov, ki ustvarijo novo Excelovo datoteko"""
    headers = make_headings()
    set_col_width(headers)
    read_txt(PATH_TXT)
    make_table()

def update_xlsx():
    """seznam ukazov, ki posodobijo obstojeco Excelovo datoteko"""
    read_txt(PATH_TXT)
    del ws_participants.tables["Udeleženci"]
    make_table()


def create_new():
    """funkcija odpre zacetno okno"""
    answer = askyesno(title="Začetno okno",
                      message="Želite ustvariti novo Excelovo datoteko?", default="no")
    return answer

def err_txt_format():
    """funkcija odpre tkinter okno z napako pri branju .txt"""
    win = tkinter.Tk()
    win.eval("tk::PlaceWindow . center")
    win.title("Napaka")
    win.geometry("500x70")

    instruct = """Tekstovne datoteke ni bilo mogoče prebrati!
Prosimo, da preverite način zapisa
(poskusite dodati/spremeniti deklaracijo statusa in leta v prvi vrstico)."""
    error = tkinter.Label(win, text=instruct, fg="red")
    error.config(font =("Couriel", 12))
    error.pack()
    win.attributes('-topmost',1)
    error.mainloop()
    sys.exit()

def err_xl_save():
    """funkcija odpre tkinter okno z napako pri shranjevanje .xlsx"""
    win = tkinter.Tk()
    win.eval("tk::PlaceWindow . center")
    win.title("Napaka")
    win.geometry("400x50")

    instruct = """Excelove datoteke ni bilo mogoče shraniti!
Poskusite jo zapreti in ponovno zagnati program."""
    error = tkinter.Label(win, text=instruct, fg="red")
    error.config(font =("Couriel", 12))
    error.pack()
    win.attributes('-topmost',1)
    error.mainloop()
    sys.exit()

def open_txt():
    """funkcija odpre .txt dokument"""
    window = tkinter.Tk()
    window.wm_attributes('-topmost', 1)
    window.withdraw()
    path_txt = tkinter.filedialog.askopenfilename(title=TXT_TITLE,
        	filetypes=txt_formats, defaultextension=".txt", initialdir=DEFAULT_DIR)
    return path_txt

def open_xlsx():
    """funkcija odpre .xlsx dokument"""
    window = tkinter.Tk()
    window.wm_attributes('-topmost', 1)
    window.withdraw()
    path_xlsx = tkinter.filedialog.askopenfilename(title=OPEN_XL_TITLE,
        filetypes=xl_formats, defaultextension=".xlsx", initialdir=DEFAULT_DIR)
    return path_xlsx

def xl_save():
    """funkcija shrani .xlsx dokument"""
    window = tkinter.Tk()
    window.wm_attributes('-topmost', 1)
    window.withdraw()
    xlsx_save = tkinter.filedialog.asksaveasfilename(title=SAVE_XL_TITLE,
        filetypes=xl_formats, defaultextension=".xlsx", initialdir=DEFAULT_DIR)
    return xlsx_save


font = Font(name= "Arial",
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000")

xl_formats=(("Ekxcelove datoteke", "*.xlsx"),)
txt_formats = (("Tekstovne datoteke", "*.txt"),)

TXT_TITLE = "Odpri tekstovno datoteko s podatki"
OPEN_XL_TITLE = "Odpri Excelovo datoteko"
SAVE_XL_TITLE = "Shrani Excelovo datoteko"
DEFAULT_DIR = "C:/Users/saša/Documents/1.SAŠA/ARPEx"


if not create_new():
    XL_LOCATION = open_xlsx()
    if XL_LOCATION != "":
        PATH_TXT = open_txt()
        database = load_workbook(XL_LOCATION)
        ws_participants = database["Udeleženci"]
        update_xlsx()
        XL_SAVE = xl_save()
else:
    PATH_TXT = open_txt()
    database = Workbook()
    ws_participants = database.active
    ws_participants.title = "Udeleženci"
    create_xlsx()
    XL_SAVE = xl_save()

try:
    database.save(XL_SAVE)
except PermissionError:
    err_xl_save()

# os.system(XL_SAVE)
os.system(f"open {XL_SAVE}") # Linux command
